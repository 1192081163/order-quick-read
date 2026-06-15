from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable, Sequence

import xlrd
from openpyxl import load_workbook

from email_order_reader.models import AttachmentParseResult, ColumnAliases, OrderRow


HEADER_SCAN_LIMIT = 20
MIN_HEURISTIC_MATCHES = 2
MIN_REASONABLE_YEAR = 2000
DEADLINE_PATTERNS = (
    r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$",
    r"^(\d{4})年(\d{1,2})月(\d{1,2})日$",
)
SheetRows = list[list[object]]


def parse_excel_attachment(
    filename: str,
    content: bytes,
    aliases: ColumnAliases | None = None,
    message_subject: str = "",
) -> AttachmentParseResult:
    aliases = aliases or ColumnAliases.default()

    try:
        sheets = _read_sheets(filename, content)
    except Exception as exc:
        return AttachmentParseResult(filename=filename, warnings=[f"{filename}：无法读取Excel附件：{exc}"])

    return _parse_sheets(filename, sheets, aliases, message_subject)


def _read_sheets(filename: str, content: bytes) -> list[SheetRows]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_openpyxl_sheets(content)
    if suffix == ".xls":
        return _read_xlrd_sheets(content)
    return []


def _read_openpyxl_sheets(content: bytes) -> list[SheetRows]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheets: list[SheetRows] = []
    try:
        for sheet in workbook.worksheets:
            rows: SheetRows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))
            sheets.append(rows)
    finally:
        workbook.close()
    return sheets


def _read_xlrd_sheets(content: bytes) -> list[SheetRows]:
    workbook = xlrd.open_workbook(file_contents=content)
    sheets: list[SheetRows] = []

    for sheet in workbook.sheets():
        rows: SheetRows = []
        for row_index in range(sheet.nrows):
            rows.append(
                [
                    _convert_xlrd_cell(
                        sheet.cell(row_index, column_index),
                        workbook.datemode,
                    )
                    for column_index in range(sheet.ncols)
                ]
            )
        sheets.append(rows)

    return sheets


def _convert_xlrd_cell(cell: xlrd.sheet.Cell, datemode: int) -> object:
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, datemode)
    return cell.value


def _parse_sheets(
    filename: str,
    sheets: Sequence[Sequence[Sequence[object]]],
    aliases: ColumnAliases,
    message_subject: str,
) -> AttachmentParseResult:
    parsed_rows: list[OrderRow] = []
    found_header = False

    for rows in sheets:
        sheet_rows = _parse_rows(filename, rows, aliases, message_subject)
        if sheet_rows is None:
            continue
        found_header = True
        parsed_rows.extend(sheet_rows)

    if not found_header:
        return AttachmentParseResult(filename=filename, warnings=[f"{filename}：未识别订单号列或截至时间列"])

    return AttachmentParseResult(filename=filename, rows=parsed_rows)


def _parse_rows(
    filename: str,
    rows: Sequence[Sequence[object]],
    aliases: ColumnAliases,
    message_subject: str,
) -> list[OrderRow] | None:
    header_match = _find_header(rows, aliases)
    if header_match is None:
        header_match = _guess_columns(rows)
    if header_match is None:
        return None

    header_index, order_col, deadline_col = header_match
    parsed_rows: list[OrderRow] = []

    for row in rows[header_index + 1 :]:
        order_number = _cell_to_text(_get_cell(row, order_col))
        deadline = _normalize_deadline(_get_cell(row, deadline_col))
        if not order_number or not deadline:
            continue

        parsed_rows.append(
            OrderRow(
                order_number=order_number,
                deadline=deadline,
                source_file=filename,
                message_subject=message_subject,
            )
        )

    return parsed_rows


def _find_header(
    rows: Sequence[Sequence[object]],
    aliases: ColumnAliases,
) -> tuple[int, int, int] | None:
    order_aliases = {_normalize_header(value) for value in aliases.order_number}
    deadline_aliases = {_normalize_header(value) for value in aliases.deadline}

    for index, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        normalized = [_normalize_header(_cell_to_text(cell)) for cell in row]
        order_col = _find_first_index(normalized, order_aliases)
        deadline_col = _find_first_index(normalized, deadline_aliases)
        if order_col is not None and deadline_col is not None:
            return index, order_col, deadline_col

    return None


def _guess_columns(rows: Sequence[Sequence[object]]) -> tuple[int, int, int] | None:
    sample_rows = rows[1:]
    if not sample_rows:
        return None

    max_columns = max((len(row) for row in sample_rows), default=0)
    best_pair: tuple[int, int] | None = None
    best_score = 0
    tied = False

    for order_col in range(max_columns):
        for deadline_col in range(max_columns):
            if order_col == deadline_col:
                continue

            paired_score = sum(
                1
                for row in sample_rows
                if _is_order_value(_get_cell(row, order_col))
                and _is_deadline_value(_get_cell(row, deadline_col))
            )

            if paired_score > best_score:
                best_pair = (order_col, deadline_col)
                best_score = paired_score
                tied = False
            elif paired_score == best_score and paired_score >= MIN_HEURISTIC_MATCHES:
                tied = True

    if best_pair is None or best_score < MIN_HEURISTIC_MATCHES or tied:
        return None

    return 0, best_pair[0], best_pair[1]


def _is_order_value(value: object) -> bool:
    text = _cell_to_text(value)
    return (
        len(text) >= 3
        and any(character.isalpha() for character in text)
        and any(character.isdigit() for character in text)
    )


def _is_deadline_value(value: object) -> bool:
    if isinstance(value, datetime | date):
        return bool(_normalize_deadline(value))

    text = _cell_to_text(value)
    if not text:
        return False

    return any(
        re.match(pattern, text) and _normalize_deadline(text)
        for pattern in DEADLINE_PATTERNS
    )


def _find_first_index(values: Iterable[str], targets: set[str]) -> int | None:
    for index, value in enumerate(values):
        if value in targets:
            return index
    return None


def _normalize_header(value: object) -> str:
    text = _cell_to_text(value).lower()
    return re.sub(r"[\s_\-:/：（）()]+", "", text)


def _get_cell(row: Sequence[object], index: int) -> object:
    if index >= len(row):
        return None
    return row[index]


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_deadline(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return _date_to_iso(value.date())
    if isinstance(value, date):
        return _date_to_iso(value)

    text = _cell_to_text(value)
    if not text:
        return ""

    for pattern in DEADLINE_PATTERNS:
        match = re.match(pattern, text)
        if match:
            year, month, day = (int(part) for part in match.groups())
            try:
                return _date_to_iso(date(year, month, day))
            except ValueError:
                return ""

    return text


def _date_to_iso(value: date) -> str:
    if value.year < MIN_REASONABLE_YEAR:
        return ""
    return value.isoformat()
