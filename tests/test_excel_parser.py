from datetime import date, datetime
from io import BytesIO

import xlwt
from openpyxl import Workbook

from email_order_reader.excel_parser import parse_excel_attachment
from email_order_reader.models import ColumnAliases


def make_xlsx(headers, rows, suffix="xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    return f"orders.{suffix}", stream.getvalue()


def make_xlsx_with_sheets(sheet_rows):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheet_rows:
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    return "orders.xlsx", stream.getvalue()


def make_xls(headers, rows):
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Orders")
    for column, value in enumerate(headers):
        sheet.write(0, column, value)
    for row_index, row in enumerate(rows, start=1):
        for column, value in enumerate(row):
            sheet.write(row_index, column, value)

    stream = BytesIO()
    workbook.save(stream)
    return "orders.xls", stream.getvalue()


def make_xls_with_native_dates(headers, rows):
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Orders")
    date_style = xlwt.XFStyle()
    date_style.num_format_str = "YYYY-MM-DD"
    time_style = xlwt.XFStyle()
    time_style.num_format_str = "HH:MM"

    for column, value in enumerate(headers):
        sheet.write(0, column, value)
    for row_index, row in enumerate(rows, start=1):
        for column, value in enumerate(row):
            if isinstance(value, tuple) and value[0] == "date":
                sheet.write(row_index, column, value[1], date_style)
            elif isinstance(value, tuple) and value[0] == "time":
                sheet.write(row_index, column, value[1], time_style)
            else:
                sheet.write(row_index, column, value)

    stream = BytesIO()
    workbook.save(stream)
    return "orders.xls", stream.getvalue()


def test_parse_xlsx_with_chinese_headers():
    filename, content = make_xlsx(
        ["订单号", "交单日期", "备注"],
        [["PO-1001", date(2026, 6, 20), "加急"]],
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.warnings == []
    assert [(row.order_number, row.deadline) for row in result.rows] == [("PO-1001", "2026-06-20")]


def test_parse_xlsm_with_english_headers_and_datetime_cell():
    filename, content = make_xlsx(
        ["Order Number", "Due Date"],
        [["PO-2002", datetime(2026, 7, 3, 9, 15)]],
        suffix="xlsm",
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.warnings == []
    assert [(row.order_number, row.deadline) for row in result.rows] == [("PO-2002", "2026-07-03")]


def test_parse_skips_empty_order_rows():
    filename, content = make_xlsx(
        ["订单编号", "截止时间"],
        [[None, date(2026, 6, 20)], ["PO-3003", "2026/06/21"]],
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.warnings == []
    assert [(row.order_number, row.deadline) for row in result.rows] == [("PO-3003", "2026-06-21")]


def test_parse_reports_missing_columns():
    filename, content = make_xlsx(
        ["客户", "金额"],
        [["ACME", 100]],
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.rows == []
    assert result.warnings == ["orders.xlsx：未识别订单号列或截至时间列"]


def test_parse_detects_headers_per_sheet_without_treating_later_headers_as_rows():
    filename, content = make_xlsx_with_sheets(
        [
            (
                "Orders A",
                [
                    ["订单号", "交单日期"],
                    ["PO-4004", date(2026, 8, 1)],
                ],
            ),
            (
                "Orders B",
                [
                    ["订单号", "交单日期"],
                    ["PO-5005", date(2026, 8, 2)],
                ],
            ),
        ]
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.warnings == []
    assert [(row.order_number, row.deadline) for row in result.rows] == [
        ("PO-4004", "2026-08-01"),
        ("PO-5005", "2026-08-02"),
    ]


def test_parse_skips_invalid_date_shaped_text_and_keeps_valid_rows():
    filename, content = make_xlsx(
        ["订单号", "交单日期"],
        [["PO-BAD", "2026/02/30"], ["PO-6006", "2026/03/01"]],
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.warnings == []
    assert [(row.order_number, row.deadline) for row in result.rows] == [("PO-6006", "2026-03-01")]


def test_parse_xls_with_chinese_headers():
    filename, content = make_xls(
        ["客户订单号", "交货日期"],
        [["PO-4004", "2026-08-09"]],
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.warnings == []
    assert [(row.order_number, row.deadline) for row in result.rows] == [("PO-4004", "2026-08-09")]


def test_parse_detects_columns_without_alias_headers():
    filename, content = make_xlsx(
        ["编号", "时间"],
        [["PO-5005", "2026-09-01"], ["PO-5006", "2026-09-02"]],
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.warnings == []
    assert [(row.order_number, row.deadline) for row in result.rows] == [
        ("PO-5005", "2026-09-01"),
        ("PO-5006", "2026-09-02"),
    ]


def test_parse_does_not_guess_unrelated_single_customer_date_row():
    filename, content = make_xlsx(
        ["Customer ID", "Last Contact"],
        [["A123", "2026-01-15"]],
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.rows == []
    assert result.warnings == ["orders.xlsx：未识别订单号列或截至时间列"]


def test_parse_xls_native_date_cell_and_rejects_time_only_cells():
    filename, content = make_xls_with_native_dates(
        ["编号", "时间"],
        [
            ["PO-7007", ("date", datetime(2026, 8, 9))],
            ["PO-TIME", ("time", datetime(1899, 12, 31, 8, 30))],
            ["PO-8008", ("date", datetime(2026, 8, 10))],
        ],
    )

    result = parse_excel_attachment(filename, content, ColumnAliases.default())

    assert result.warnings == []
    assert [(row.order_number, row.deadline) for row in result.rows] == [
        ("PO-7007", "2026-08-09"),
        ("PO-8008", "2026-08-10"),
    ]
